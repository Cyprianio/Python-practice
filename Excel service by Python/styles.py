from openpyxl.utils import get_column_letter
from openpyxl.styles import Color, PatternFill

def cell_naming(ws):
    ws['H1'] = 'Dofinansowanie łącznie:'
    ws['H2'] = 'Wkład własny łącznie:'
    ws['H3'] = 'Koszt całkowity:'

    ws['H11'] = 'Kategoria 1:'
    ws['H12'] = 'Kategoria 5:'
    ws['H13'] = 'Kategoria 6:'
    ws['H14'] = 'Suma kategorii:'
    ws['H15'] = 'Koszty pośr 5%:'

    ws['H18'] = 'Kategoria 1:'
    ws['H19'] = 'Kategoria 5:'
    ws['H20'] = 'Kategoria 6:'
    ws['H21'] = 'Suma kategorii:'
    ws['H22'] = 'Koszty pośr 5%:'
    ws['H23'] = 'Suma całk pośr:'


    ws['I1'] = '=SUM(C2:C10)'
    ws['I2'] = '=SUM(D2:D10)'
    ws['I3'] = '=SUM(E2:E10)'

    ws['I9'] = 'prace rozwojowe'
    ws['I10'] = 'Koszty całkowite'
    ws['I11'] = """=SUMIFS($E$2:$E$10,$A$2:$A$10,"prace rozwojowe",$B$2:$B$10,1)"""
    ws['I12'] = """=SUMIFS($E$2:$E$10,$A$2:$A$10,"prace rozwojowe",$B$2:$B$10,5)"""
    ws['I13'] = """=SUMIFS($E$2:$E$10,$A$2:$A$10,"prace rozwojowe",$B$2:$B$10,6)"""
    ws['I14'] = '=SUM(I11:I13)'
    ws['I15'] = '=PRODUCT(0.05*I14)'
    ws['I16'] = 'badania przemysłowe'
    ws['I17'] = 'Koszty całkowite'
    ws['I18'] = '=SUMIFS($E$2:$E$10,$A$2:$A$10,"prace przemyslowe",$B$2:$B$10,1)'
    ws['I19'] = '=SUMIFS($E$2:$E$10,$A$2:$A$10,"prace przemyslowe",$B$2:$B$10,5)'
    ws['I20'] = '=SUMIFS($E$2:$E$10,$A$2:$A$10,"prace przemyslowe",$B$2:$B$10,6)'
    ws['I21'] = '=SUM(I18:I20)'
    ws['I22'] = '=PRODUCT(0.05*I21)'
    ws['I23'] = '=I15+I22'

    ws['J1'] = "Dofinansowanie + koszty pośrednie:"
    ws['J2'] = "Wkład własny + koszty pośrednie:"
    ws['J3'] = "Koszt całk + koszty pośrednie:"

    ws['J10'] = 'Dofinansowanie:'
    ws['J11'] = '=SUMIFS($C$2:$C$10,$A$2:$A$10,"prace rozwojowe",$B$2:$B$10,1)'
    ws['J12'] = '=SUMIFS($C$2:$C$10,$A$2:$A$10,"prace rozwojowe",$B$2:$B$10,5)'
    ws['J13'] = '=SUMIFS($C$2:$C$10,$A$2:$A$10,"prace rozwojowe",$B$2:$B$10,6)'
    ws['J14'] = '=SUM(J11:J13)'
    ws['J15'] = '=PRODUCT(0.05*J14)'

    ws['J17'] = 'Dofinansowanie:'
    ws['J18'] = '=SUMIFS($C$2:$C$10,$A$2:$A$10,"prace przemyslowe",$B$2:$B$10,1)'
    ws['J19'] = '=SUMIFS($C$2:$C$10,$A$2:$A$10,"prace przemyslowe",$B$2:$B$10,5)'
    ws['J20'] = '=SUMIFS($C$2:$C$10,$A$2:$A$10,"prace przemyslowe",$B$2:$B$10,6)'
    ws['J21'] = '=SUM(J18:J20)'
    ws['J22'] = '=PRODUCT(0.05*J21)'
    ws['J23'] = '=J15+J22'

    ws['K1'] = '=SUM(I1+J23)'
    ws['K2'] = '=SUM(I2+K23)'
    ws['K3'] = '=SUM(I3+I23)'

    ws['K10'] = 'Wkład własny:'
    ws['K11'] = '=I11-J11'
    ws['K12'] = '=I12-J12'
    ws['K13'] = '=I13-J13'
    ws['K14'] = '=SUM(K11:K13)'
    ws['K15'] = '=PRODUCT(0.05*K14)'

    ws['K17'] = 'Wkład własny:'
    ws['K18'] = '=I18-J18'
    ws['K19'] = '=I19-J19'
    ws['K20'] = '=I20-J20'
    ws['K21'] = '=SUM(K18:K20)'
    ws['K22'] = '=PRODUCT(0.05*K21)'
    ws['K23'] = '=K15+K22'
    return ws

def coloring(ws):
    blueFill = PatternFill('solid', fgColor='6064DA')

    yellowFill = PatternFill('solid', fgColor='FFCF0A')                

    orangeFill = PatternFill('solid', fgColor='FF8D1A')

    for col in range(1, 6):
        ws[get_column_letter(col) + '1'].fill = blueFill

    for row in range(1, 4):
        ws['I' + str(row)].fill = yellowFill
        ws['K' + str(row)].fill = yellowFill

    for col in range(9, 12):
        ws[get_column_letter(col) + '15'].fill = yellowFill
        ws[get_column_letter(col) + '22'].fill = yellowFill
        ws[get_column_letter(col) + '23'].fill = yellowFill

    ws['I9'].fill = orangeFill
    ws['I16'].fill = orangeFill
    return ws


def stretch_cells(ws):
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 5
    ws.column_dimensions['G'].width = 5
    ws.column_dimensions['H'].width = 25
    ws.column_dimensions['I'].width = 25
    ws.column_dimensions['J'].width = 35
    ws.column_dimensions['K'].width = 20
    
    return ws

def format_cells(ws):
    for row in range(2, 11):
        ws['C' + str(row)].number_format = '#,##0.00'
        ws['D' + str(row)].number_format = '#,##0.00'
        ws['E' + str(row)].number_format = '#,##0.00'

    for row in range(1, 4):
        ws['I' + str(row)].number_format = '#,##0.00'
        ws['K' + str(row)].number_format = '#,##0.00'

    for row in range(11, 16):
        ws['I' + str(row)].number_format = '#,##0.00'
        ws['J' + str(row)].number_format = '#,##0.00'
        ws['K' + str(row)].number_format = '#,##0.00'

    for row in range(18, 24):
        ws['I' + str(row)].number_format = '#,##0.00'
        ws['J' + str(row)].number_format = '#,##0.00'
        ws['K' + str(row)].number_format = '#,##0.00'
