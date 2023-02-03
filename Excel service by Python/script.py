import pandas as pd
from openpyxl import Workbook, load_workbook
from styles import cell_naming, coloring, stretch_cells, format_cells

data = pd.read_excel('plik_zrodlowy_zalacznik1.xlsx', sheet_name=0)

df_data = pd.DataFrame(data)
df_data.drop('NrZadania', axis=1, inplace=True)
df_data.drop('PartnerOdpowiedzialny', axis=1, inplace=True)
df_data.drop('NazwaKosztu', axis=1, inplace=True)

excelWriter = pd.ExcelWriter('BUDZET_SZCZEGOLOWY.xlsx')
df_data.to_excel(excelWriter, index=False)
excelWriter.save()

wb = Workbook()
wb = load_workbook('BUDZET_SZCZEGOLOWY.xlsx')
ws = wb.active

cell_naming(ws)
coloring(ws)
stretch_cells(ws)
format_cells(ws)

wb.save('BUDZET_SZCZEGOLOWY.xlsx')
